# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

S = json.load(open("data.json", encoding="utf-8"))
ALG = {"G":"Gluten","M":"Süt","Y":"Yumurta","B":"Balık","K":"Kabuklu deniz ürünleri",
       "S":"Soya","SS":"Susam","N":"Sert kabuklu yemişler","F":"Yer fıstığı",
       "H":"Hardal","C":"Kereviz","SO":"Sülfit","MO":"Yumuşakçalar","L":"Acı bakla"}
CODES = list(ALG.keys())

INK="FF21201A"; MOSS="FF425C2B"; BRASS="FF8A6113"
wb = Workbook(); ws = wb.active; ws.title = "Menü"

hdr = ["Bölüm","Grup","Ürün (TR)","Ürün (EN)","Fiyat ₺","İçindekiler (TR)","İçindekiler (EN)"] + CODES + ["Notunuz"]
ws.append(hdr)

thin = Side(style="thin", color="FFD5D2C4")
for c in range(1, len(hdr)+1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=MOSS)
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
ws.row_dimensions[1].height = 40

r = 2
for s in S:
    for g in s["g"]:
        for it in g["i"]:
            row = [s["tr"], g["tr"], it["n"], it["ne"], it["p"],
                   it.get("ing",""), it.get("inge","")]
            a = set(it.get("a", []))
            row += ["X" if c in a else "" for c in CODES]
            row += [""]
            ws.append(row)
            for c in range(1, len(hdr)+1):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(name="Arial", size=10)
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=(c in (6,7)))
            ws.cell(row=r, column=5).number_format = '#,##0'
            ws.cell(row=r, column=5).font = Font(name="Arial", size=10, bold=True, color=BRASS)
            ws.cell(row=r, column=5).alignment = Alignment(vertical="top", horizontal="right")
            for c in range(8, 8+len(CODES)):
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.font = Font(name="Arial", size=10, bold=True, color=BRASS)
                cell.fill = PatternFill("solid", fgColor="FFFFF9E8")
            ws.cell(row=r, column=len(hdr)).fill = PatternFill("solid", fgColor="FFFFFFCC")
            r += 1
last = r-1

widths = [15,24,32,32,9,46,46] + [5]*len(CODES) + [26]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{last}"

dv = DataValidation(type="list", formula1='"X,"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)
dv.add(f"H2:{get_column_letter(7+len(CODES))}{last}")

# --- Nasıl kullanılır ---
ws2 = wb.create_sheet("Nasıl kullanılır")
lines = [
 ("KÖK CAFE — Alerjen ve içerik kontrol listesi", True),
 ("", False),
 ("Bu dosyadaki alerjen işaretleri ve içindekiler bilgileri TASLAKTIR.", False),
 ("Ürün adlarından ve verdiğiniz malzeme listelerinden çıkarılmıştır; mutfak ekibinizce", False),
 ("kontrol edilmeden menüde yayınlanmamalıdır.", False),
 ("", False),
 ("Nasıl düzeltirsiniz:", True),
 ("1. \"Menü\" sekmesinde her satır bir üründür.", False),
 ("2. Alerjen sütunlarında (G, M, Y, ...) ürün o alerjeni içeriyorsa hücreye X yazın,", False),
 ("   içermiyorsa hücreyi boş bırakın. Sarı hücreler doldurulacak alanlardır.", False),
 ("3. İçindekiler sütunlarını serbestçe düzeltebilir, boş olanları doldurabilirsiniz.", False),
 ("4. Fiyat değişikliklerini de bu dosyada yapabilirsiniz.", False),
 ("5. Dosyayı bana geri gönderin; menüyü güncellerim. QR kod değişmez.", False),
 ("", False),
 ("Örnek satır:", True),
 ("Bölüm: Tatlılar | Grup: Tatlılar | Ürün: Künefe | Fiyat: 420", False),
 ("İçindekiler: Kadayıf, dil peyniri, tereyağı, şerbet, Antep fıstığı", False),
 ("G: X   M: X   N: X   (gluten, süt ve sert kabuklu yemiş içerir)", False),
 ("", False),
 ("Alerjen kodları:", True),
]
for code, name in ALG.items():
    lines.append((f"   {code}  =  {name}", False))
for i,(txt,bold) in enumerate(lines, start=1):
    c = ws2.cell(row=i, column=1, value=txt)
    c.font = Font(name="Arial", size=11, bold=bold, color=MOSS if bold else INK)
ws2.column_dimensions["A"].width = 95

wb.save("KOK-alerjen-kontrol.xlsx")
print("satır:", last-1, "| sütun:", len(hdr))
